from django.contrib import messages
import openpyxl
import pyodbc
from openpyxl.styles import PatternFill, Font
import random
import time

INSERT_INCREMENTO_META = """INSERT INTO HOMOLOGACAO.DBO.ML_12062023_TESTE_METAINCREMENTO(
									COMP,
									CARTEIRA,
									FRENTE,
									TIPO_META,
									META_UM,
									META_DOIS,
									META_TRES,
									META_QUATRO,
									META_CINCO
								)
VALUES"""
QUERY_METATIVA_INCREMENTO_META = """SELECT 
    TOP 5 
	* 
FROM
	HOMOLOGACAO.DBO.ML_12062023_TESTE_METAINCREMENTO
WHERE 
"""
ALTERANDO_METATIVA_INCREMENTO_META ="""UPDATE HOMOLOGACAO.DBO.ML_12062023_TESTE_METAINCREMENTO
	SET ATIVA = 'INATIVO'
WHERE  
"""

QUERY_META_OPERADORES = """
SELECT 
	*
FROM 
	HOMOLOGACAO.DBO.ML_07072023_TESTE_METASOPERADORES
WHERE 
	COMPETENCIA    =  '$VARIAVEL_COMPETENCIA$' AND
	NOME_CREDOR  =  '$VARIAVEL_NOME_CREDOR$'
"""


def gerador_nome():

    def random_number():
        return (random.randint(1,100))
    
    nome =  f'-{random_number()}'
    return nome

def desativando(cursor,cn):
    cursor.close()
    cn.close()

def verify_carteira(carteira):
    """
    True: Existe
    False: Não Existe

    Essa função verifica se a carteira 
    que foi enviada existe"""

    #TODO: Seria interresante puxar as carteiras existentes, de cada meta, de um banco de dados
    # Assim, sempre que uma nova carteira surgir, é so adicionar no banco de dados e ela passará a passar na nossa validação
    # Por enquanto se uma nova carteira surge, temos que mudar no hardcode

    # True - Existe   |   False - Não existe
    carteira_existe = False

    carteiras_existentes = (
        'alto_ticket',
        'comercial_imobiliario',
        'veiculos_amg',
        'veiculos_lp',
        'consorcio_imo_jur',
        'comercial_juri',
        'credito_imobiliario_juamg',
        'veiculos_juri',
    )

    if carteira in carteiras_existentes:
            carteira_existe = True

    return carteira_existe

def is_excel(plan):
        """
        True: é uma planilha excel
        False: Não é uma planilha excel
        """
        
        try: # Só abre se for excel, caso não vai dar erro
            openpyxl.load_workbook(plan)
            is_excel = True
        except:
            is_excel =  False

        return is_excel

def sheet_exists(plan):
        """
        A página de dados existe?
        Sim : True
        Não : False

        Essa função verifica se a aba de dados existe, para que possamos ter acesso aos dados.
        """

        if 'Dados' not in plan.sheetnames:
            exist = False
        else:
            exist = True

        return exist

def headers_is_corretct(plan,tipo):
        """
        True: Está correto
        False: Está incorreto

        Essa função verifica se os cabeçalhos da meta existem, caso não existam, possívelmente estamos abrindo
        uma planilha errada, que possui uma aba nomeada DADOS.
        """
        header_plan = []
        headers_default = {
            'incremento' : [
                'COMP',
                'CARTEIRA',
                'FRENTE', 
                'TIPO_META', 
                'META_1', 
                'META_2', 
                'META_3', 
                'META_4',
                'META_5'
            ],
            'operadores' : [
                'COMPETENCIA',	
                'DATA_IMPORT',	
                'QUEM_IMPORTOU',	
                'COD_CRED',	
                'NOME_CREDOR',	
                'CENTRO_CUSTO',	
                'COD_FUNC',	
                'CPF_FUNC',	
                'NOME_FUNCIONARIO',	'SUPERVISOR',	
                'FRENTE',	
                'META_QTDE',	 
                'META_HONORARIOS', 	 
                'META_REPASSE', 	 
                'META_VALOR', 	 
                'META_ATIVA', 	
                'TURNO',	
                'ATUAÇÃO',	
                'ESTAGIO',	
                'DATA_INI',	
                'DATA_FIN',	
                'TIPO_MEDICAO',
            ],
            'deflatores' : [
                'COMP',	
                'CENTRO_CUSTO',	
                'COD_CRED',	
                'NOM_CREDOR',	
                'COD_FUNC',	
                'NOME_OPERADOR',	
                'CPF',	
                'MATRICULA',
                'FALTAS_QTDE',	
                'ADVERTENCIAS_QTDE',
                'RECLAMAÇÕES_QTDE',	
                'MONITORIA_NOTA',
                'SUSPENSAO_QTDE',	
                'QUEM_VALIDOU',	
                'DATA_VALIDAÇÃO'
            ],
            'premio'     : [
                'DATA_IMPORT',	
                'COMP',	
                'COD_CRED',	
                'CARTEIRA',	
                'FRENTE', 
                'TIPO_META',	
                'TIPO_MEDICAO',	
                'CRITERIO_1',	
                'VLR_PREMIO_1',	
                'CRITERIO_2',	
                'VLR_PREMIO_2',	
                'CRITERIO_3',	
                'VLR_PREMIO_3',	
                'CRITERIO_4',	
                'VLR_PREMIO_4',	
                'CRITERIO_5',	
                'VLR_PREMIO_5',	
                'CRITERIO_6',	
                'VLR_PREMIO_6',	
                'CRITERIO_7',	
                'VLR_PREMIO_7',	
                'CRITERIO_8',	
                'VLR_PREMIO_8'
                ],
        }
        max_cols = {
            'incremento' : 9,
            'operadores' : 22,
            'deflatores' : 15,
            'premio'     : 23,
        }
        
        header_default = headers_default[tipo]
        

        max_col = max_cols[tipo]
        


        #TODO: Criar um sistema que muda o local de busca de valores, já que em metas operadores está em um local e incremento meta em outro
        for row in plan.iter_rows(min_row=3, max_row=3, max_col=max_col ):
            for cell in row:
                header_plan.append(cell.value)

        return header_plan == header_default

def get_data_meta_com_duplicatas(plan : object, max_col : int):
    """
    ## Recebe
    plan    : página com dados da meta
    max_col : coluna máxima que devemos considerar

    ## Retorno
    Uma lista, com listas. Em cada lista filha há os valores, separados por virgula, de uma linha
        [
            [valor1,valor2,valor3],
            [valor1,valor2,valor3],
            [valor1,valor2,valor3],
            
        ]

    ### Função
    Recuperar todos os dados da meta, contendo duplicatas.

    """

    # Lista pai
    # [  ]
    dados = []

    # Passa por cada coluna
    for row in plan.iter_rows(min_row=4, max_col=max_col):

        # Adiciona uma lista a lista pai
        # [ [1,2,3], [nova...] ]
        dados.append([])

        # Adicionando valores na lista criada acima
        for cell in row:

            # [ [1,2,3], [adicionando...]]
            dados[-1].append(cell.value)  
    
    # [ [1,2,3], [1,2,3] ]
    return dados

def recuperar_informacoes_unicas(plan : object , max_col : int):
    """
    ## Parâmetros
    plan    -> *página* de dados, a página que contém as informações da meta.
    max_col -> a última coluna que possui dados na planilha.

    ## Retorno
    Dicionário.
    {
        'COMPETÊNCIA' : set(),

        'NOME_CRED' : set(),
    }

    
    ### Explicação
        Recebemos a planilha e a última coluna que devemos considerar os valores, 
    começamos a iterar cada coluna e recuperar seus valores. 
    
    -- Durante o looping
        Pegamos o primeiro valor da coluna e criamos uma key no diconário e damos um set como
    valor da key. Então criamos outro looping que itera cada célula da coluna.

        -- Durante o loopings
            Caso o valor da célula seja igual a key do dicionário, que é o nome da coluna,
            pulamos o looping.
            -- pular looping

            Entramos em cada célula da coluna e adicionamos no dicinário de dados, na key com
        o nome da coluna, que é um set. Inserimos todas as células.
        -- Fim looping

    -- Fim looping

    Então retornamos um dicionário, cada key do dicionário é o nome de uma coluna, cada key possui
    um set, que por sua vez, possui os valores da coluna.

    """

    dados = {

    }

    # Passa por cada coluna
    for row in plan.iter_cols(min_row=3, max_col=max_col):

        # Criando uma key com o primeiro valor da coluna, que é o nome da coluna
        dados[row[0].value] = set()

        # Adicionando valores na key atual
        for cell in row:
            # Se o valor da célula for o nome da coluna, pule o looping
            if cell.value == row[0].value:
                continue

            dados[row[0].value].add(cell.value)  

    return dados

def format_number(plan):

    fundo_vermelho = PatternFill("solid", fgColor="EB607F")
    font_bold = Font(bold=True)

    for coluna in plan.iter_cols(min_col=5 ,min_row=4, max_col=9):
        for cell in coluna:
            if type(cell.value) != int:
                cell.fill = fundo_vermelho
                cell.font = font_bold
    return plan

def pegue_um_numero_aleatorio():
    numro_um = random.randint(0,999)
    numro_dois = random.randint(1,9)
    numro_tres = random.randint(9,99)

    return f'{numro_um}{numro_dois}{numro_tres}'

def execute_consulta(consulta):
    
    def conectando():

        server = '10.10.5.30'
        database = 'METAS'
        username='planejamento'
        password= 'pl@n1234'
        #cn = pyodbc.connect('DRIVER={"SQL Server"};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cn = pyodbc.connect(f'DRIVER={"SQL Server"};SERVER=10.10.5.30;DATABASE=HOMOLOGACAO;UID=planejamento;PWD=pl@n1234')
        
        cursor = cn.cursor()
        return (cursor,cn)
                
    try:
        cursor,cn = conectando()
    except:
        return None, None, True
    # Executando uma consulta
    try:
        cursor.execute(consulta)
    except:
        #return cursor.fetchone, cursor,cn, True    
        return cursor, cn, True    
            
    # Pegando o resultado da consulta
    #return cursor.fetchone, cursor,cn, False
    return cursor, cn, False

def informacoes_corretas(
        request: str ,
        #carteira: str, 
        meta:str,
        meta_arquivo: object,
        meta_op=None):
        
        """ 
        ## Recebimento
        request         : str       -> A request que o client nos enviou
        carteira        : str       -> A carteira que capturamos no front-end
        meta_op         : object    -> O woorkbook excel, arquivo, da meta de operadores
        incremento_meta : object    -> O woorkbook excel, arquivo, da meta incremento

        
        
        #### Retorno -- Correto
        Uma tupla 
        CASA 0 - Valor False 
        CASA 1 - Uma tupla com a página de dados de cada planilha aberta.
        
        (False, (   incremento_meta_dados,  meta_op_dados)   )
        
        Caso correta retornams false, para não entrar na função que retornari uma renderização.
        E uma tupla com as sheets, para continuarmos o trabalho com os dados corretos.

        #### Retorno -- Falso
        Uma tupla
        CASA 0 - Request
        CASA 1 - Template

        Caso errada, retornamos uma request com template pois utilizaremos na função render.

        

        ### Código da função
        CODE_FUNCTION : A0


        ### Explicação
            Essa função tem o objetivo de verificar a existência da carteira, meta_op e incremento_meta.
        Saber se o usuário enviou ou não. Caso não tenha enviado, como precisamos dessa informação, deve
        ->mos retornar um erro, ào usuário, informando que ele deve enviar essas informações.

            Sabendo que essas informações existem, precisamos saber se elas estão corretas. Então vali
        ->damos. Caso errada retornamos um erro informando o erro para o usuário, assim ele pode
        corrigir.
        """

        # Os arquivos enviados, são realmentes planilhas excel?
        """if meta != 'operadores':
            if not is_excel(meta_op): 
                messages.add_message(request,messages.ERROR, "Ops! O que foi enviado no campo meta operadores, não era uma planilha ")
                return (request, 'planilhas/metas.html')
            else:
                meta_op = openpyxl.load_workbook(meta_op) # Abrindo a planilha
            """
        if not is_excel(meta_arquivo):
            messages.add_message(request,messages.ERROR, "Ops! O que foi enviado no campo incremento meta, não era uma planilha ")
            return (request, 'planilhas/metas.html')
        else:
            meta_arquivo = openpyxl.load_workbook(meta_arquivo)



        # Verificando se a página de dados existe
        """if meta != 'operadores':
            if not sheet_exists(meta_op):
                messages.add_message(
                    request,
                    messages.ERROR, 
                    "Ops! a página 'dados' na planilha meta operadores, não pode ser encontrada. Baixe nosso layout de metas e insira os dados!")
                return (request, 'planilhas/metas.html')
            else:
                meta_op_dados = meta_op['Dados']
            """
        if not sheet_exists(meta_arquivo):
            messages.add_message(
                request,
                messages.ERROR, 
                "Ops! a página 'dados' na planilha, não pode ser encontrada. Baixe nosso layout de metas e insira os dados!")
            return (request, 'planilhas/metas.html')
        else:
            meta_pagina_dados = meta_arquivo['Dados']


        
        # Verificando se os cabecalhos são os esperados
        if not headers_is_corretct(meta_pagina_dados,meta):
            messages.add_message(
                request,
                messages.WARNING,
                "Não foi possível encontrar o cabeçalho de informações, na aba de dados. Você está enviando a meta correta? Verifique!"
            )
            return (request, 'planilhas/metas.html')

        """ if meta != 'operadores':
            if not headers_is_corretct(meta_op_dados,'operadores'):
                messages.add_message(
                    request,
                    messages.WARNING,
                    "Não foi possível encontrar o cabeçalho de informações, na aba de dados, na planilha meta operadores. Você está enviando a meta correta? Verifique!"
                )
                return (request, 'planilhas/metas.html')
        """
        """if meta != 'operadores':
            return (False, (meta_pagina_dados, meta_op_dados))"""
        return (False, (meta_pagina_dados, None))

def comparacao_dados(request : str,dados_incremento : dict ,dados_operadores : dict):

    """
    ## Recebimento
    request             : a request que recebemos do cliente
    dados_incremento    : dicionário com os dados da meta de incrementos
    dados_operadores    : dicionário com os dados do ranking de operadores

    ## Retorno
    Apenas retornamos algo se houver um erro nas informações.
    Retornamos dados necessários para renderizar a página com o erro.
    """

    competencia_incorreta = False if  len(dados_incremento['COMP']) == 1 else True
    if competencia_incorreta:
        messages.add_message(request,messages.WARNING, "Há um erro na coluna de competências")
        return (request, 'planilhas/metas.html')   
    

    for frente in dados_incremento['FRENTE']:
        if frente not in dados_operadores['FRENTE']:
            frente_incorreta_dois = True
            break
        else:
            frente_incorreta_dois = False
    
    for frente in  dados_operadores['FRENTE']:
        if frente not in dados_incremento['FRENTE']:
            frente_incorreta = True
            break
        else:
            frente_incorreta = False    

    if frente_incorreta or frente_incorreta_dois:
        messages.add_message(request,messages.WARNING, "Há um erro na coluna de frentes da planilha meta incremento ou a meta de operadores, deste mês e desta carteira, ainda não foi importada ou você não preencheu todas as frentes que estão presentes na meta de operadores")
        return (request, 'planilhas/metas.html')          

    for tipo_meta in dados_incremento['TIPO_META']:
        if tipo_meta not in dados_operadores['TIPO_MEDICAO']:
            tipometa_incorreta = True
            break
        else:
            tipometa_incorreta = False

    if tipometa_incorreta:
        messages.add_message(request,messages.WARNING, "O tipo de meta está difente da planilha de meta operadores")
        return (request, 'planilhas/metas.html') 

def validando_nome_carteira(request:str, carteira: str, dados_da_planilha : list):
    # palavras_chaves = {
    #     'alto_ticket':   [
    #         'alt',
    #         'tick'
    #         ],

    #     'comercial_imobiliario':  [
    #         'com,imo','rid'
    #         ],

    #     'veiculos_amg': [
    #         'vei','amig'
    #         ],

    #     'veiculos_lp':[
    #         'vei','lp'
    #         ],

    #     'consorcio_imo_jur':  [
    #         'consor','imo','jur'
    #         ],

    #     'comercial_juri': [
    #         'com','jur'
    #         ],

    #     'credito_imobiliario_juamg':  [
    #         'cred','imo',(
    #     'ju','amg'
    #     )
    #         ],

    #     'veiculos_juri': [
    #         'vei',
    #         'jur'
    #     ]
    # }
    
    #caracteres_chaves = palavras_chaves[carteira] # Pegando os caracteres chaves do nome da carteira.
    
    while True:
        carteira = carteira.replace('-',' ')

        if not '-' in carteira:
            break
    caracteres_chaves = carteira.split()

    for nome_carteira in dados_da_planilha['CARTEIRA']:
        # pegando cada nome de carteira existente na planilha

        int_na_nome_carteira = True if type(nome_carteira) == int else False
        if int_na_nome_carteira:
            messages.add_message(
                request,
                    messages.ERROR,
                    'Há um valor numérico na coluna de nome da carteira, corrija!'
                    )
            return True
    
        nome_carteira = nome_carteira.lower()
        for caractere in caracteres_chaves:
            # pegando cada caracter chave para comparar

            if not caractere in nome_carteira:
                messages.add_message(
                    request,
                        messages.WARNING,
                        'O nome da carteira contido na planilha é diferente da carteira informada por você no campo acima! '
                        )
                return True
    
    return False   
                    
def formatando_planilha(request,sheet_metaincremento,dados_incremento, incremento_meta_woorkbook):
    row = {
        'META_1' : {
            'coluna_dados' : dados_incremento['META_1'],
            'coluna_local' : 5
            },

        'META_2' : {
            'coluna_dados' : dados_incremento['META_2'],
            'coluna_local' : 6
            },

        'META_3' : {
            'coluna_dados' : dados_incremento['META_3'],
            'coluna_local' : 7
            },

        'META_4' : {
            'coluna_dados' : dados_incremento['META_4'],
            'coluna_local' : 8
            },

        'META_5' : {
            'coluna_dados' : dados_incremento['META_5'],
            'coluna_local' : 9
            },
    }     
    
    exists_str = bool()

    # Passando em cada letra, de cada meta e verificando se há letras ou espaços
    for conjunto_dados in row.values():

        for meta in conjunto_dados['coluna_dados']:

            if type(meta) != int:
                exists_str = True
                break
        
    if exists_str:
        format_number(sheet_metaincremento) # Formatando campos incorretos da página.
        messages.add_message(request,messages.ERROR,'Alguns dados contidos na planilhas, não são do tipo correto')

        while True:
            try:
                numero_aleatorio = pegue_um_numero_aleatorio()
                incremento_meta_woorkbook.save(
                    #f'templates/global/planilhas/planilha_com_erro{numero_aleatorio}.xlsx'
                    f'asdas{numero_aleatorio}.xlsx'
                    )
                nome_planilha_incorreta = f'planilha_com_erro{numero_aleatorio}.xlsx'
    
                break
            except:
                ...

        contexto = {
            'url' : f'/media/{nome_planilha_incorreta}'
        }

        return(
            request, 
            'planilhas/metas.html',
            contexto
        ) 

def existe_valores_incorretos(colunas : tuple):
            
            """
            ### Recebimento
            Recebemos uma tupla com os dados das colunas de metas,
            (
                valor_um,
                valor_dois,
                valor_tres,
            )

            ### Retorno
            Caso haja um valor incorreto, retornamos True
            """

            for coluna in colunas:
    
                for celula in coluna:
                    if type(celula) != int:
                        return True

def criando_valors_para_insert(valores,coluna_maxima):
            """
            Essa função cria a consulta que irá inserir valores no banco de dados.

            """

            consulta = str()


            linhas_existentes_planilha =len(valores)
            linha_atual = 0
            fim_consulta = False
            for row in valores:
                consulta += ' ( '

                celula_atual = 0
                for cell in row:

                    # Adicionando o valor na consulta
                    if cell == None:
                        fim_consulta = True
                        continue
                    if type(cell) == str or "datetime.datetime" in str(type(cell)):
                        if "datetime.datetime" in str(type(cell)):
                            cell = str(cell)
                        consulta += "'"
                        consulta += cell
                        consulta += "'"
                    else:
                        consulta += str(cell)
                    
                    celula_atual += 1
                    if celula_atual < coluna_maxima:
                        consulta += ', '
                
                linha_atual += 1
                if linha_atual < linhas_existentes_planilha and not fim_consulta:
                    consulta += ' ), '
                elif fim_consulta:
                    consulta = consulta[0:-5]
                    return consulta
                else:
                    consulta += ') '
            return consulta

def contar_linhas(pagina):
    linhas = int()
    linhas_desconsideradas = 3

    for linha in pagina.iter_rows(min_row=4,min_col=2, max_col=2):

        for celula in linha:
            linhas += 1
    return linhas + linhas_desconsideradas  
    
def padronizando_nome_cartira(
        incremento_meta_dados,
        #carteira,
        nome_padrao_carteira,
        total_linhas):
    coluna_nomecredor = incremento_meta_dados['B']
    coluna_tipo_meta = incremento_meta_dados['D']


    # nomes_padrao_carteiras = {
    # 'alto_ticket'   :  'ALTO_TICKET_PADRAO',

    # 'comercial_imobiliario':  'COMERCIAL_IMOBILIARIO_PADRAO',

    # 'veiculos_amg': 'VEICULOS_AMIGAVEL_PADRAO',

    # 'veiculos_lp': 'VEICULOS_LP_PADRAO',

    # 'consorcio_imo_jur': 'CONSORCIO_IMO_JURIDICO',

    # 'comercial_juri': 'CONSORCIO_JURIDICO' ,

    # 'credito_imobiliario_juamg':  'CREDITO_IMOBILIARIO_JUAMG',
    #     }
    
    nome_padrao = nome_padrao_carteira
    
    linha = 1
    for celula_nome_carteira in coluna_nomecredor:
        
        nome_carteira = nome_padrao

        # Caso já esteja na linha 4, linha que começa a possuir dados
        if linha >=4:   
            celula_nome_carteira.value = nome_carteira

        linha += 1
    
def criando_query_meta_ativa(carteiras,competencia,consulta_base):
            consulta = consulta_base 
            
            for comp in competencia:
                consulta += f" COMP = '{comp}' AND "
                

            total_carteiras = len(carteiras) #-- 3
            carteira_adicionadas_consulta = 1
            
            for carteira in carteiras:
                
                # Primeiro valor adicionado e tem que haver mais de uma carteira
                if carteira_adicionadas_consulta == 1 and total_carteiras > 1:
                    consulta += ' ( '

                # Não é a última carteira
                if carteira_adicionadas_consulta < total_carteiras:
                    consulta += f"CARTEIRA = '{carteira}' OR   "

                # É a última carteira
                elif carteira_adicionadas_consulta == total_carteiras:
                    consulta  += f"CARTEIRA = '{carteira}'   "

                # É a última carteira, mas existem várias carteiras
                if carteira_adicionadas_consulta == total_carteiras and total_carteiras >1:
                    consulta += ' ) ' 
                if carteira_adicionadas_consulta == total_carteiras:
                    consulta += " AND ATIVA = 'ATIVA' "

                carteira_adicionadas_consulta += 1

            return consulta

def existem_linhas_excedentes(pagina,max_col):
            
    linhas_em_cada_coluna = set()
    
    for coluna in pagina.iter_cols(min_row=4,max_col=max_col):
        quantidade_linhas = 0

        for celula in coluna:
            

            celulaa_preenchida = True if celula.value != None else False
            if celulaa_preenchida:
                quantidade_linhas += 1

        linhas_em_cada_coluna.add(quantidade_linhas)

    linhas_excedentes = True if len(linhas_em_cada_coluna) > 1 else False

    return linhas_excedentes
