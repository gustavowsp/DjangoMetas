# Utils utilizados
import planilhas.viewsfunc.view_incremento as incremento_functions
import planilhas.viewsfunc.view_operador as utils_op
import planilhas.viewsfunc.view_deflatores as utils_deflatores
import planilhas.viewsfunc.view_valorpremio as utils_premio

# Models utilizados
from planilhas.models import Carteira

from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.contrib import messages

# Libs utilizadas
import openpyxl
from datetime import datetime
import time

def formatar_carteira(carteira):
    """
    Essa função recebe uma string que contêm duas informações
    - Cod cred da carteira
    - Nome carteira
    Vamos formatar a carteira de forma que peguemos essas duas informações separadamente
    """
    ...

    carteira = carteira.split(';')
    return carteira

def formatar_cpf(cpf):
    
    tamanho_cpf = len(cpf)
    if tamanho_cpf == 10:
        cpf = '0' + cpf
    elif tamanho_cpf == 9:
        cpf = '00' + cpf
    elif tamanho_cpf == 8:
        cpf = '000' + cpf
    elif tamanho_cpf == 7:
        cpf = '0000' + cpf

    return cpf


# Create your views here.
def index(request):
    return render(request, 'planilhas/index.html')

def layouts(request):
    return render(request,'planilhas/layouts.html')

def incremento(request):

    if not request.user.is_authenticated:
        messages.add_message(request,messages.INFO,"É necessário estar autenticado para começar o processo de envio de metas!!")
        return redirect('pagina_usuario:login')
    
    todas_carteiras = Carteira.objects.filter(ativa=True)

    context = {
        'carteiras' :  todas_carteiras
    }

    if request.method == 'POST':
        

        carteira = request.POST.get('carteira')

        if carteira:   
            carteira = formatar_carteira(carteira)
            cod_cred,nome_carteira = carteira[0], carteira[1]
        else:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de selecionar uma carteira"
                )
            return render(request, 'planilhas/metas.html',context)
    

        try:
            carteira_informacoes_object = Carteira.objects.get(cod_cred=cod_cred,ativa=True,nome_cred_padrao=nome_carteira) 
        except:
            messages.add_message(request,messages.ERROR, 'A carteira informada está inativa ou é inexistente. Caso exista, contate-nos para adicionarmos esta carteira')
            return render(request, 'planilhas/metas.html',context)

        # C000
        try:
            #meta_op = request.FILES['meta_mae']
            incremento_meta = request.FILES['meta_filha']
            # Pegamos as duas planilhas
        except:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de enviar uma das planilhas"
                )
            return render(request, 'planilhas/metas.html',context)
            # O usuário não enviou as planilhas e recuperamos algo que não existe, ocasionando em um erro.
        
        # Verificamos que as planilhas de metas foram enviadas.


        # C001
        metas_erradas, planilhas_or_template,  = incremento_functions.informacoes_corretas(
            request,
            #carteira=carteira,
            #meta_op=meta_op,
            meta_arquivo=incremento_meta,
            meta='incremento'
        )
        

        # Verificamos se as planilhas estão corretas e se a carteira existe.

        # R000
        if metas_erradas:
            return render(
                metas_erradas, # Existe um requeste nesta variável.
                planilhas_or_template,context
            )
        # As metas estavam erradas então vamos retornar um erro

        incremento_meta = openpyxl.load_workbook(incremento_meta)



        # C002
        # Pegando a sheet de dentro da tupla, que foi retornada na função anteior.
        incremento_sheet_dados = planilhas_or_template[0]
        metaop_sheet_dados = planilhas_or_template[1]
        incremento_sheet_dados = incremento_meta['Dados']
        # C003
        # Verificando se há linhas excedentes na planilha
        existem_linhas_excedentes = incremento_functions.existem_linhas_excedentes(incremento_sheet_dados,max_col=9)
        #existem_linhas_excedentes_dois = incremento_functions.existem_linhas_excedentes(metaop_sheet_dados)
        if existem_linhas_excedentes: #or existem_linhas_excedentes_dois:
            messages.add_message(request,messages.ERROR, 'Em sua planilha há linhas que possuem células vazias, preenchas!')
            return render(request,'planilhas/metas.html',context)

        for linha in incremento_sheet_dados.iter_rows(max_col=9,min_row=4,max_row=4):
            for celula in linha:
                if None == celula.value:
                    messages.add_message(request,messages.ERROR, 'Você enviou uma planilha totalmente vazia, por favor preencha!')
                    return render(request,'planilhas/metas.html',context)
        # Formatando o tipo de medicao
        for tipo_medicao in incremento_sheet_dados.iter_rows(min_row=4,max_col=4, min_col=4):

            valor_celula = tipo_medicao[0].value
            if valor_celula == None:
                continue

            valor_celula = valor_celula.replace('-','_')
            
            while ' ' in valor_celula:
               
                valor_celula = valor_celula.replace(' ','')

            tipo_medicao[0].value = valor_celula
        
        # Pegando os valores das planilhas.
        dados_incremento = incremento_functions.recuperar_informacoes_unicas(incremento_sheet_dados,9)
        
        
        competencia = list(dados_incremento.get('COMP'))[0]
        nome_credor = carteira_informacoes_object.nome_cred_padrao

        consulta = incremento_functions.QUERY_META_OPERADORES
        consulta = consulta.replace('$VARIAVEL_COMPETENCIA$',competencia).replace('$VARIAVEL_NOME_CREDOR$',nome_credor)
        itens = incremento_functions.execute_consulta(consulta)
        query_falhou = itens[2]

        if query_falhou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados",
                {
            'carteiras':todas_carteiras
            }
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            meta_existente = itens[0].fetchall()
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql

        dados_teste_operador = {
            'COMPETENCIA' : set() ,
            'DATA_IMPORT' : set() ,
            'QUEM_IMPORTOU' : set() ,
            'COD_CRED' : set() ,
            'NOME_CREDOR' : set() ,
            'CENTRO_CUSTO' : set() ,
            'COD_FUNC' : set() ,
            'CPF_FUNC' : set() ,
            'NOME_FUNCIONARIO' : set() ,
            'SUPERVISOR' : set() ,
            'FRENTE' : set() ,
            'META_QTDE' : set() ,
            'META_HONORARIOS' : set() ,
            'META_REPASSE' : set() ,
            'META_VALOR' : set() ,
            'META_ATIVA' : set() ,
            'TURNO' : set() ,
            'ATUACAO' : set() ,
            'ESTAGIO' : set() ,
            'DATA_INI' : set() ,
            'DATA_FIN' : set() ,
            'TIPO_MEDICAO' : set() ,
        }
        for linha in meta_existente:
            linha_atual = 0

            for value in dados_teste_operador.values():
                value.add(linha[linha_atual])
                linha_atual += 1

        dados_operadores = dados_teste_operador

       
        for tipo_medicao in incremento_sheet_dados.iter_rows(min_row=4,max_col=4, min_col=4):
            print(tipo_medicao[0].value)

        # C004
        # Comparando os valores de cada coluna chave com a planilha de metas operadores
        metas_erradas = incremento_functions.comparacao_dados(request,dados_incremento,dados_operadores)
        if metas_erradas:
            request = metas_erradas[0]
            template = metas_erradas[1]
            return render(request,template,context)

        # Verificando se o nome está correto
        palavras_chaves = carteira_informacoes_object.palavras_chaves
        nome_carteira_errado = incremento_functions.validando_nome_carteira(request,palavras_chaves,dados_incremento)
        if nome_carteira_errado:
            return render(request, 'planilhas/metas.html',context)
        
        # Alterando o nome_cred na planilha para o nome padrão
        total_linhas_pagina = incremento_functions.contar_linhas(incremento_sheet_dados)
        incremento_sheet_dados = incremento_meta['Dados']  
        nome_padrao_credor = carteira_informacoes_object.nome_cred_padrao      
        incremento_functions.padronizando_nome_cartira(incremento_sheet_dados,nome_padrao_credor,total_linhas_pagina)
        
        # Pegando os novos valores, já que alterei acima.
        dados_incremento = incremento_functions.recuperar_informacoes_unicas(incremento_sheet_dados,9)


        # C005 -- Formatando células com valores diferentes de int
        row = (
            dados_incremento['META_1'],
            dados_incremento['META_2'],
            dados_incremento['META_3'],
            dados_incremento['META_4'],
            dados_incremento['META_5'],
            )   
        # Há células com outro valor a não ser int?
        if incremento_functions.existe_valores_incorretos(row):
            
            # Formatando as células incorretas
            incremento_functions.format_number(incremento_sheet_dados) 
            
            name = incremento_functions.gerador_nome()
            path = 'media/planilhas/sua_planilha-'
            incremento_meta.save(f'{path}{name}.xlsx')

            context['planilha_incorreta_url'] =f'{path}{name}.xlsx'
            context['path'] = 'http://127.0.0.1:8000/'
            
            #     'planilha_incorreta_url' : f'{path}{name}.xlsx',
            #     'path' : 'http://127.0.0.1:8000/'
            # }

            messages.add_message(
                request,
                messages.ERROR,
                'Alguns dados contidos na planilhas, não são do tipo correto')
            return render(
                request, 
                'planilhas/metas.html',
                context
            ) 


        carteiras_da_meta = dados_incremento['CARTEIRA']
        competencia_da_meta = dados_incremento['COMP']
       
        puxando_meta_ativa = incremento_functions.criando_query_meta_ativa(
            carteiras_da_meta,
            competencia_da_meta,
            incremento_functions.QUERY_METATIVA_INCREMENTO_META)

        itens = incremento_functions.execute_consulta(puxando_meta_ativa)
        query_falhou = itens[2]

        if query_falhou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            meta_existente = itens[0].fetchall()
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql

        # Alterar meta antiga para inativa
        if meta_existente:
            alterando_meta_ativa = incremento_functions.criando_query_meta_ativa(
                carteiras_da_meta,
                competencia_da_meta,
                incremento_functions.ALTERANDO_METATIVA_INCREMENTO_META
                )
            itens = incremento_functions.execute_consulta(alterando_meta_ativa)
            
            importacao_nao_funcionou = itens[2] 
            if importacao_nao_funcionou:
                incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
                messages.add_message(
                    request,messages.
                    ERROR,
                    "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos alterar meta antiga no banco de dados"
                    )
                return render(request, 'planilhas/metas.html',context)
            else:
                itens[1].commit() # Confirmando as alterações    
                incremento_functions.desativando(itens[0],itens[1]) # Desativando 



        # C006 - Importando a meta -- Como ativa
        data_planilha_incremento = incremento_functions.get_data_meta_com_duplicatas(incremento_sheet_dados,9)# Pegando os valores da planilha
        consulta = incremento_functions.criando_valors_para_insert(data_planilha_incremento,9)# Criando a consulta -- Parte dos VALUES
        consulta_final = (incremento_functions.INSERT_INCREMENTO_META + consulta)# Juntando a consulta de VALUES com insert -- COMANDO INTO
        itens = incremento_functions.execute_consulta(consulta_final)# Importando para o banco de dados
        
        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados",
                {
            'carteiras':todas_carteiras
            }
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            itens[1].commit() # Confirmando as alterações    
            incremento_functions.desativando(itens[0],itens[1]) # Desativando


        
        # Deu tudo certo e estamos importando
        messages.add_message(request,messages.SUCCESS, "Importada com sucesso!")
        return render(request, 'planilhas/metas.html',context)

    return render(request, 'planilhas/metas.html',context)

def valor(request):
    
    
    if not request.user.is_authenticated:
        messages.add_message(request,messages.INFO,"É necessário estar autenticado para começar o processo de envio de metas!!")
        return redirect('pagina_usuario:login')
    
    context = {
        'carteiras' : Carteira.objects.filter(ativa=True)
    }

    if request.method == 'POST':
    
        carteira = request.POST.get('carteira')

        if carteira:   
            carteira = formatar_carteira(carteira)
            cod_cred,nome_carteira = carteira[0], carteira[1]
        else:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de selecionar uma carteira"
                )
            return render(request, 'planilhas/metas.html',context)
    
        
        # C000
        try:
            meta_op = request.FILES['meta_filha']
            # Pegamos as duas planilhas
        except:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de enviar uma das planilhas"
                )
            return render(request, 'planilhas/metas.html',context)
            # O usuário não enviou as planilhas e recuperamos algo que não existe, ocasionando em um erro.
        

        # C001
        metas_erradas, planilhas_or_template,  = incremento_functions.informacoes_corretas(
            request,
            #carteira=carteira,
            meta_arquivo=meta_op,
            meta='premio'
        )
        # Verificamos se as planilhas estão corretas e se a carteira existe.

        # R000
        if metas_erradas:
            return render(
                metas_erradas, # Existe um requeste nesta variável.
                planilhas_or_template,
                context
            )
        # As metas estavam erradas então vamos retornar um erro



        # C002
        # Pegando a sheet de dentro da tupla, que foi retornada na função anteior.
        meta_sheet_dados = planilhas_or_template[0]
        meta_op = openpyxl.load_workbook(meta_op)
        meta_sheet_dados = meta_op['Dados']
    
        # C003
        # Verificando se há linhas excedentes na planilha
        #existem_linhas_excedentes = incremento_functions.existem_linhas_excedentes(meta_sheet_dados)
        existem_linhas_excedentes = incremento_functions.existem_linhas_excedentes(meta_sheet_dados,max_col=23)
        if existem_linhas_excedentes:
            messages.add_message(request,messages.ERROR, 'Em sua planilha há linhas que possuem células vazias, preenchas!')
            return render(request,'planilhas/metas.html',context)
        
        
        for linha in meta_sheet_dados.iter_rows(max_col=23,min_row=4,max_row=4):
            for celula in linha:
                if None == celula.value:
                    messages.add_message(request,messages.ERROR, 'Você enviou uma planilha totalmente vazia, por favor preencha!')
                    return render(request,'planilhas/metas.html',context)
                
        ### COMEÇO ALGORITIMO

        competencias = set()
        for linha_da_coll_comp in meta_sheet_dados.iter_rows(min_col=2,max_col=2, min_row=4):
            
            valor_celula = linha_da_coll_comp[0].value 

            if not valor_celula:
                break

            competencia = valor_celula
            if len(competencia) != 7:
                messages.add_message(request, messages.WARNING, 'Comptência está no formatato errado, o correto é "2023-01" ')
                return render(request, 'planilhas/metas.html',context)
            comeco,meio,fim  = competencia [-2:],competencia[4:5],competencia[0:4]

            comptencia_formato_correto = True if comeco.isdigit() and '-' in meio and fim.isdigit() else False
            if not comptencia_formato_correto:
                messages.add_message(request, messages.WARNING, 'Os valores contidos na competência estão incorretos')
                return render(request, 'planilhas/metas.html',context)

            competencias.add(valor_celula)
        if len(competencias) > 1:
            messages.add_message(request, messages.WARNING, 'Existem mais de uma competência em sua planilha. Apenas uma competência é permitida')
            return render(request, 'planilhas/metas.html',context)


        try:
            informacao_carteira_object = Carteira.objects.get(cod_cred=cod_cred,nome_cred_padrao=nome_carteira,ativa=True) 
        except:
            messages.add_message(request,messages.ERROR, 'A carteira informada está inativa ou é inexistente. Caso exista, contate-nos para adicionarmos esta carteira')
            return render(request, 'planilhas/metas.html',context)
        
        cod_cred_carteira           = informacao_carteira_object.cod_cred
        nome_padrao_carteira        = informacao_carteira_object.nome_cred_padrao
        palavras_chaves_carteira    =  informacao_carteira_object.palavras_chaves
        while True:
            
            palavras_chaves_carteira = palavras_chaves_carteira.replace('-',' ')
            
            if not '-' in palavras_chaves_carteira:
                palavras_chaves_carteira = palavras_chaves_carteira.split()
                break
            
        # Validando o cod_cred da planilha
        for celula in meta_sheet_dados.iter_rows(min_row=4, min_col=3, max_col=3):
            
            valor_celula = str(celula[0].value).strip()
            if valor_celula == 'None':
                break

            time.sleep(5)
            if not valor_celula ==  cod_cred_carteira:

                messages.add_message(request,messages.ERROR, 'Você selecionou a carteira incorreta... O cod cred selecionado é diferente da planiha')
                return render(request, 'planilhas/metas.html',context)

        # Validando nome da carteira
        for celula in meta_sheet_dados.iter_rows(min_row=4, min_col=4, max_col=4):
            valor_celula = celula[0].value
            valor_celula = str(valor_celula).strip().lower()

            if valor_celula == 'none':
                break

            for palavra in palavras_chaves_carteira:
                
                if palavra not in valor_celula:
                    messages.add_message(request,messages.ERROR, 'O nome do credor existente na planilha está incorreto, verifique!')
                    return render(request, 'planilhas/metas.html',context)

            celula[0].value = nome_padrao_carteira

        # Inserindo a data atual na
        data_atual = str(datetime.now())
        data_atual = data_atual[:-3]

        for linha in meta_sheet_dados.iter_rows(min_row=4, min_col=1, max_col=1):
            valor_celula = linha[0].value

            if valor_celula == None:
                break
            linha[0].value = data_atual

        
        # Iterando a coluna C-3 COD CRED
        for celula in meta_sheet_dados.iter_rows(min_row=4,min_col=3,max_col=3):
            
            tipo_medicao = celula[0].value
            tipo_medicao = str(tipo_medicao)

            if tipo_medicao == 'None':
                break
            
            if not tipo_medicao == cod_cred_carteira:
                messages.add_message(request,messages.WARNING,'O cod cred presente na planilha não corresponde a carteira informada')
                return render(request,'planilhas/metas.html',context)

        
        # Iterando a coluna D-4  NOME_ CREDOR
        for celula in meta_sheet_dados.iter_rows(min_row=4,min_col=4,max_col=4):
            tipo_medicao = celula[0].value
            nome_credor = str(tipo_medicao).strip().lower()

            if nome_credor == 'none':
                break

            for palavra_chave in palavras_chaves_carteira:
                
                if palavra_chave not in nome_credor:
                    messages.add_message(request,messages.WARNING,'O nome credor da planilha está incorreto...')
                    return render(request,'planilhas/metas.html',context)

        # Criando consulta para recuperar informações da meta de operadores
        competencia_da_meta = list(competencias)[0]
        query_meta_op = utils_premio.QUERY_META
        query_meta_op += f"COMPETENCIA = '{competencia_da_meta}' AND\n NOME_CREDOR = '{nome_padrao_carteira}'"
        
        itens =  incremento_functions.execute_consulta(query_meta_op)

        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            if itens[0]:
                incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
                messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            consulta_meta_operadores = itens[0].fetchall() # Confirmando as alterações    
            incremento_functions.desativando(itens[0],itens[1]) # Desativando

        # Criando locais para armazenar os resultados da Query 
        frente_mt_operadors,tipo_meta_operador = set(), set()

        # Inserindo os valores nos sets acima
        for meta in consulta_meta_operadores:
            frente_mt_operadors.add(meta[0])
            tipo_meta_operador.add(meta[1])
        
        # Validando as frentes
        frentes_plan = set()
        for celula in meta_sheet_dados.iter_rows(min_row=4,max_col=5,min_col=5):
            
            frente = celula[0].value
            
            if frente == None:
                break

            if frente not in frente_mt_operadors:
                messages.add_message(request,messages.WARNING,'Existem frentes que não correspondem às da meta de operadors em sua planilha')
                return render(request,'planilhas/metas.html',context)
            frentes_plan.add(frente)

        for frente in frente_mt_operadors:
            if frente not in frentes_plan:
                messages.add_message(request,messages.WARNING,'Existem frentes da meta de operadores que não estão inseridas na planilha enviada.')
                return render(request,'planilhas/metas.html',context)                        
       
       # Validando o tipo de medicao 
        for celula in meta_sheet_dados.iter_rows(min_row=4,max_col=6,min_col=6):    
            tipo_medicao = celula[0].value
            if tipo_medicao == None:
                break

            tipo_medicao_formato_um = tipo_medicao.replace(' - ','_')
            celula[0].value = tipo_medicao_formato_um
            
            if tipo_medicao not in tipo_meta_operador and tipo_medicao_formato_um not in tipo_meta_operador:
                messages.add_message(request,messages.WARNING,'Existem tipo de medicao que não correspondem às da meta de operadors em sua planilha')
                return render(request,'planilhas/metas.html',context)


        coluna_a_ser_validada = 8
        criterios_anteriores = list()
        for vez in range(0,8):
            
            linha_atual = 0

            for linha_de_dados in meta_sheet_dados.iter_rows(min_row=4,min_col=coluna_a_ser_validada,max_col=coluna_a_ser_validada + 1):
                criterio        =   linha_de_dados[0].value
                valor_premio    =   linha_de_dados[1].value

                if criterio == None or valor_premio == None:
                    break

                # Checando se critério e valor prêmio são números
                if not type(criterio) == int and not type(criterio) == float:
                    messages.add_message(request,messages.WARNING,'O critério da planilha deve ser um número')
                    return render(request,'planilhas/metas.html',context)     
                               
                if not type(valor_premio) == int and not type(valor_premio) == float :
                    messages.add_message(request,messages.WARNING,'O valor prêmio da planilha deve ser um número')
                    return render(request,'planilhas/metas.html',context)     

                if coluna_a_ser_validada == 8:
                    
                    if criterio != 100.01:
                        messages.add_message(request,messages.WARNING,'O valor do primeiro critério deve ser 100.01')
                        return render(request,'planilhas/metas.html',context)  
                    


                    criterios_anteriores.append(criterio)

                    continue

                if criterio == 0:

                    if valor_premio != 0:
                        messages.add_message(request,messages.WARNING,'Você não preencheu o critério, logo o valor do prêmio não deve estar preenchido!')
                        return render(request,'planilhas/metas.html',context)
                    
                    continue


                criteiro_anterior = criterios_anteriores[linha_atual]

                if criteiro_anterior < criterio:
                    criterios_anteriores.pop(linha_atual)
                    criterios_anteriores.insert(linha_atual,criterio)
                else:   
                    messages.add_message(request,messages.INFO,f'Critério antigo:{criteiro_anterior} critério atual: {criterio}')
                    messages.add_message(request,messages.WARNING,'Os critérios devem sempre ser maiores que o seu antecessor. Atualmente há'
                                         ' um critério que é menor que anteior')
                    return render(request,'planilhas/metas.html',context)  
                linha_atual += 1
            coluna_a_ser_validada += 2

        
        dados = incremento_functions.get_data_meta_com_duplicatas(meta_sheet_dados,max_col=23)
        print(dados)
        print('#'*50) 
        valores_insert = incremento_functions.criando_valors_para_insert(dados,23)
        insert_consulta = utils_premio.INSERT_META
        consulta = insert_consulta + valores_insert


        itens = incremento_functions.execute_consulta(consulta)
        print(consulta)
        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            itens[0].commit() # Confirmando as alterações
            incremento_functions.desativando(itens[0],itens[1]) # Desativando


        messages.add_message(request,messages.SUCCESS,'Meta importada com sucesso')
        

    return render(request,'planilhas/metas.html',context)

def deflator(request):  
    
    context = {
        'carteiras' : Carteira.objects.filter(ativa=True)
    }
    
    if not request.user.is_authenticated:
        messages.add_message(request,messages.INFO,"É necessário estar autenticado para começar o processo de envio de metas!!")
        return redirect('pagina_usuario:login')
    
    if request.method == 'POST':
    
        carteira = request.POST.get('carteira')

        if carteira:   
            carteira = formatar_carteira(carteira)
            cod_cred,nome_carteira = carteira[0], carteira[1]
        else:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de selecionar uma carteira"
                )
            return render(request, 'planilhas/metas.html',context)
    
        # C000
        try:
            meta_op = request.FILES['meta_filha']
            # Pegamos as duas planilhas
        except:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de enviar uma das planilhas"
                )
            return render(request, 'planilhas/metas.html',context) 
            # O usuário não enviou as planilhas e recuperamos algo que não existe, ocasionando em um erro.
        
        # C001
        metas_erradas, planilhas_or_template,  = incremento_functions.informacoes_corretas(
            request,
            #carteira=carteira,
            meta_arquivo=meta_op,
            meta='deflatores'
        )
        # Verificamos se as planilhas estão corretas e se a carteira existe.

        # R000
        if metas_erradas:
            return render(
                metas_erradas, # Existe um requeste nesta variável.
                planilhas_or_template ,
                context
            )
        # As metas estavam erradas então vamos retornar um erro

        # C002
        # Pegando a sheet de dentro da tupla, que foi retornada na função anteior.
        meta_sheet_dados = planilhas_or_template[0]
        meta_op = openpyxl.load_workbook(meta_op)
        meta_sheet_dados = meta_op['Dados']
    
        # C003
        # Verificando se há linhas excedentes na planilha
        existem_linhas_excedentes = incremento_functions.existem_linhas_excedentes(meta_sheet_dados,max_col=15)
        if existem_linhas_excedentes:
            messages.add_message(request,messages.ERROR, 'Em sua planilha há linhas que possuem células vazias, preenchas!')
            return render(request,'planilhas/metas.html',context)
        
        for linha in meta_sheet_dados.iter_rows(max_col=15,min_row=4,max_row=4):
            for celula in linha:
                if None == celula.value:
                    messages.add_message(request,messages.ERROR, 'Você enviou uma planilha totalmente vazia, por favor preencha!')
                    return render(request,'planilhas/metas.html',context)
                   
        ############## COMEÇO ALGORITMO
        competencias = set()
        for linha_da_coll_comp in meta_sheet_dados.iter_rows(min_col=1,max_col=1, min_row=4):
            
            valor_celula = linha_da_coll_comp[0].value 

            if not valor_celula:
                break

            competencia = valor_celula
            if len(competencia) != 7:
                messages.add_message(request, messages.WARNING, 'Comptência está no formatato errado, o correto é "2023-01" ')
                return render(request, 'planilhas/metas.html',context)
            comeco,meio,fim  = competencia [-2:],competencia[4:5],competencia[0:4]

            comptencia_formato_correto = True if comeco.isdigit() and '-' in meio and fim.isdigit() else False
            if not comptencia_formato_correto:
                messages.add_message(request, messages.WARNING, 'Os valores contidos na competência estão incorretos')
                return render(request, 'planilhas/metas.html',context)

            competencias.add(valor_celula)
        if len(competencias) > 1:
            messages.add_message(request, messages.WARNING, 'Existem mais de uma competência em sua planilha. Apenas uma competência é permitida')
            return render(request, 'planilhas/metas.html',context)

        try:
            informacao_carteira_object = Carteira.objects.get(cod_cred=cod_cred,nome_cred_padrao=nome_carteira,ativa=True) 
        except:
            messages.add_message(request,messages.ERROR, 'A carteira informada está inativa ou é inexistente. Caso exista, contate-nos para adicionarmos esta carteira')
            return render(request, 'planilhas/metas.html',context)
        
        cod_cred_carteira           = informacao_carteira_object.cod_cred
        nome_padrao_carteira        = informacao_carteira_object.nome_cred_padrao
        palavras_chaves_carteira    =  informacao_carteira_object.palavras_chaves
        while True:
            palavras_chaves_carteira = palavras_chaves_carteira.replace('-',' ')
            if not '-' in palavras_chaves_carteira:
                palavras_chaves_carteira = palavras_chaves_carteira.split()
                break

        # Inserindo centro de custos  
        for celula in meta_sheet_dados.iter_rows(min_row=4,min_col=2,max_col=2):
            valor_cell = celula[0].value
            if valor_cell == None:
                break
            celula[0].value =  informacao_carteira_object.centro_custo

        # Validando cod cred
        for celula in meta_sheet_dados.iter_rows(min_row=4, min_col=3, max_col=3):
            valor_celula = str(celula[0].value).strip()

            if valor_celula == 'None':
                break

            if not valor_celula ==  cod_cred_carteira:

                messages.add_message(request,messages.ERROR, 'Você selecionou a carteira incorreta... O cod cred selecionado é diferente da planiha')
                return render(request, 'planilhas/metas.html',context)

        # Validando o nome do credor
        for celula in meta_sheet_dados.iter_rows(min_row=4, min_col=4, max_col=4):
            valor_celula = celula[0].value

            if not valor_celula:
                break

            valor_celula = str(valor_celula).strip().lower()

            for palavra in palavras_chaves_carteira:
                
                if palavra not in valor_celula:
                    messages.add_message(request,messages.ERROR, 'O nome do credor existente na planilha está incorreto, verifique!')
                    return render(request, 'planilhas/metas.html',context)

            celula[0].value = nome_padrao_carteira

        # Validando código Funcionário
        query_script =  utils_deflatores.CODIGO_FUNCIONARIO

        # Executando consulta
        itens = incremento_functions.execute_consulta(query_script)

        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            query = itens[0].fetchall() # Confirmando as alterações
            incremento_functions.desativando(itens[0],itens[1]) # Desativando

        # Formatando query de códigos de func
        range_query = len(query)
        for resultado in range(range_query):
            
            # Formatando conjuntos de dados da query
            cod_funcionario = str(query[resultado][0]).strip()
            cpf_funcionario = str(query[resultado][1]).strip()
            
            # Inserindo conjunto de dados formatados
            query[resultado][0] = cod_funcionario
            query[resultado][1] = cpf_funcionario

        # Entrando em cada código funcionário da planilha
        for celula in meta_sheet_dados.iter_rows(min_row=4,min_col=5,max_col=7):

            cod_func = str(celula[0].value)
            cpf_func = str(celula[2].value)
            cpf_func = formatar_cpf(cpf_func)

            # Buscando o código do funcionário 
            for conjunto_dados in query:
                
                # Pegando o código de funcionário existente no banco de dados
                cod_func_query = conjunto_dados[0]

                # Verificando se o cod func da planilha existe no banco de dados
                if cod_func == cod_func_query:
                    cod_func_correto = True
                    break

                else:
                    cod_func_correto = False
            
            if not cod_func_correto:
                
                # Buscando o cpf do funcionário
                for conjunto_dados in query:
                    
                    # Pegando o cpf de funcionário existente no banco de dados
                    cod_func_query = conjunto_dados[0]
                    cpf_func_query = conjunto_dados[1]

                    # Verificando se o cpf func da planilha existe no banco de dados
                    if cpf_func == cpf_func_query:
                        celula[0].value = True

                        # Inserindo o codigo funcionário correto
                        celula[0].value = cod_func_query
                        break


        # Validando valores de monitoria
        for linha in meta_sheet_dados.iter_rows(min_row=4,min_col=9, max_col=13):
            
            # Verificando se os valores da linha podem se tornar INT:
            for valor in linha:
                
                # Caso seja None já validamos tudo que era necessário.
                if valor.value == None:
                    break

                try:
                    int(valor.value)
                except:
                    messages.add_message(request,messages.WARNING, 'Os valores da coluna I até  M devem ser numéricos não caracteres.')
                    return render(request, 'planilhas/metas.html',context)
            
        lista_cpfs = set()
        dados_funcionario = {

        }

        # Adicionando cpfs na lista de cpfs 
        for looping_h in meta_sheet_dados.iter_rows(min_row=4, max_col=7, min_col=7):
            cpf = str(looping_h[0].value).strip().lower()
            
            if cpf == 'none':
                break

            tamanho_cpf = len(cpf)
            if tamanho_cpf == 10:
                cpf = '0' + cpf
                looping_h[0].value = cpf
            elif tamanho_cpf == 9:
                cpf = '00' + cpf
                looping_h[0].value = cpf
            elif tamanho_cpf == 8:
                cpf = '000' + cpf
                looping_h[0].value = cpf
            elif tamanho_cpf == 7:
                cpf = '0000' + cpf
                looping_h[0].value = cpf
                
            lista_cpfs.add(cpf)

            dados_funcionario[cpf] = {
                'cpfs_query'                :   '',
                'nome_func_query'           :   '' ,
                'matricula_func_query'      :   '', 
            }

        consulta = utils_deflatores.QUERY_FUNCIONARIOS
        cpfs_iterados = 1

        # Passando cada cpf para a consulta, assim formando uma query
        for cpf in lista_cpfs:
            
            if cpfs_iterados == len(lista_cpfs):
                consulta += f"'{cpf}'"
            else:
                consulta += f"'{cpf}',"

            cpfs_iterados += 1
        
        consulta += ')'

        itens =  incremento_functions.execute_consulta(consulta)
        
        # Verificando se a query funcionou, caso funcione vamos recuperar as informações da query
        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            if itens[0]:
                incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
                messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            cpfs = itens[0].fetchall() # Confirmando as alterações    
            incremento_functions.desativando(itens[0],itens[1]) # Desativando
        
        for dados_func in cpfs:

            cpf_    = dados_func[1]
            nome_   = dados_func[0]
            codigo_ = dados_func[2] 

            dados_funcionario[cpf_]['cpfs_query'] = cpf_
            dados_funcionario[cpf_]['nome_func_query']  = nome_
            dados_funcionario[cpf_]['matricula_func_query'] = codigo_
        
        # Validando dados do funcionário
        for linha in meta_sheet_dados.iter_rows(min_row=4, min_col=5, max_col=8):
            nome_func = str(linha[1].value).strip().lower()
            cpf_func=   str(linha[2].value).strip().lower()

            if nome_func == 'none' or cpf_func == 'none':
                break

            dados_dos_funcionarios =    dados_funcionario[cpf_func]
            cpf_func_query         =    str(dados_dos_funcionarios['cpfs_query']).strip().lower()
            nome_func_query        =    str(dados_dos_funcionarios['nome_func_query']).strip().lower() 


            if not cpf_func in cpf_func_query:
                messages.add_message(request,messages.ERROR, 'Alguns nomes de funcionários estão incorretos.')
                return render(request,'planilhas/metas.html',context)

            linha[3].value = dados_dos_funcionarios['matricula_func_query']
        data_atual = str(datetime.now())
        data_atual = data_atual[:-3]
        usuario = request.user.username

        for linha in meta_sheet_dados.iter_rows(min_row=4, min_col=14, max_col=15):
            valor_cell_zero = linha[0].value
            valor_cell_um   = linha[1].value

            if valor_cell_zero == None or valor_cell_um == None:
                break
            linha[0].value = usuario
            linha[1].value = data_atual
    
        dados = incremento_functions.get_data_meta_com_duplicatas(meta_sheet_dados,max_col=15)
        dados_arrumados = list()
        for valor in dados:
            if valor[0] == None:
                break

            valor.append(usuario) # Adicionando quem importou como p nultimo
            valor.append(data_atual) # Adicionando a data da importação como último
            dados_arrumados.append(valor)


            # Adicionando os campos DATA IMPORT E QUEM_IMPORTOU VIA PYTHON
            # Essses campos não existem na planilha, já na tabela existem.
            # O valor é igual à um campo já existente na planilha, então apenas vamos copiar e colar
            
        valores_insert = incremento_functions.criando_valors_para_insert(dados_arrumados,17)
        insert_consulta = utils_deflatores.INSERT_META
        consulta = insert_consulta + valores_insert

        itens = incremento_functions.execute_consulta(consulta)
        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            try:
                query_meta = itens[0].commit() # Confirmando as alterações    
            except:
                incremento_functions.desativando(itens[0],itens[1]) # Desativando

        messages.add_message(request,messages.SUCCESS, 'Meta importada com sucesso')
    
    return render(request, 'planilhas/metas.html',context)     

def meta_operador(request):
        
    if not request.user.is_authenticated:
        messages.add_message(request,messages.INFO,"É necessário estar autenticado para começar o processo de envio de metas!!")
        return redirect('pagina_usuario:login')

    todas_carteiras = Carteira.objects.filter(ativa=True)
    context = {
        'carteiras' : todas_carteiras
    }

    if request.method == 'POST':
    
        carteira = request.POST.get('carteira')
        

        if carteira:   
            carteira = formatar_carteira(carteira)
            cod_cred,nome_carteira = carteira[0], carteira[1]
        else:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de selecionar uma carteira"
                )
            return render(request, 'planilhas/metas.html',context)
    
        # C000
        try:
            meta_op = request.FILES['meta_filha']
            # Pegamos as duas planilhas
        except:
            messages.add_message(
                request,
                messages.WARNING, 
                "Ops! Você se esqueceu de enviar uma das planilhas"
                )
            return render(request, 'planilhas/metas.html',context)
            # O usuário não enviou as planilhas e recuperamos algo que não existe, ocasionando em um erro.
        

        # C001
        metas_erradas, planilhas_or_template,  = incremento_functions.informacoes_corretas(
            request,
            #carteira=carteira,
            meta_arquivo=meta_op,
            meta='operadores'
        )
        # Verificamos se as planilhas estão corretas e se a carteira existe.

        # R000
        if metas_erradas:
            return render(
                metas_erradas, # Existe um request nesta variável.
                planilhas_or_template,context)
        # As metas estavam erradas então vamos retornar um erro



        # C002
        # Pegando a sheet de dentro da tupla, que foi retornada na função anteior.
        meta_sheet_dados = planilhas_or_template[0]
        meta_op = openpyxl.load_workbook(meta_op)
        meta_sheet_dados = meta_op['Dados']



        # C002
        # Pegando a sheet de dentro da tupla, que foi retornada na função anteior.
        #meta_sheet_dados = planilhas_or_template[0]
    
    
        # C003
        # Verificando se há linhas excedentes na planilha
        existem_linhas_excedentes = incremento_functions.existem_linhas_excedentes(meta_sheet_dados,max_col=22)
        if existem_linhas_excedentes:
            messages.add_message(request,messages.ERROR, 'Em sua planilha há linhas que possuem células vazias, preenchas!')
            return render(request,'planilhas/metas.html',context)
        
        for linha in meta_sheet_dados.iter_rows(max_col=22,min_row=4,max_row=4):
            for celula in linha:
                if None == celula.value:
                    messages.add_message(request,messages.ERROR, 'Você enviou uma planilha totalmente vazia, por favor preencha!')
                    return render(request,'planilhas/metas.html',context)


        ############## COMEÇO ALGORITMO
        competencias = set()
        for linha_da_coll_comp in meta_sheet_dados.iter_rows(min_col=1,max_col=1, min_row=4):
            
            valor_celula = linha_da_coll_comp[0].value 

            if not valor_celula:
                break

            competencia = valor_celula
            if len(competencia) != 7:
                messages.add_message(request, messages.WARNING, 'Comptência está no formatato errado, o correto é "2023-01" ')
                return render(request, 'planilhas/metas.html',context)
            comeco,meio,fim  = competencia [-2:],competencia[4:5],competencia[0:4]

            comptencia_formato_correto = True if comeco.isdigit() and '-' in meio and fim.isdigit() else False
            if not comptencia_formato_correto:
                messages.add_message(request, messages.WARNING, 'Os valores contidos na competência estão incorretos')
                return render(request, 'planilhas/metas.html',context)

            competencias.add(valor_celula)
        if len(competencias) > 1:
            messages.add_message(request, messages.WARNING, 'Existem mais de uma competência em sua planilha. Apenas uma competência é permitida')
            return render(request, 'planilhas/metas.html',context)


        data_atual = str(datetime.now())
        data_atual = data_atual[:-3]

        for linha_da_coll_dataimport in meta_sheet_dados.iter_rows(min_col=2,max_col=2, min_row=4):
            
            if  linha_da_coll_dataimport[0].value: 
                linha_da_coll_dataimport[0].value = data_atual

        #TODO: Recuperar qual usuário esta importando
        usuario_nome = request.user.username

        for linha_da_coll_quemimporta in meta_sheet_dados.iter_rows(min_col=3,max_col=3, min_row=4):
            if linha_da_coll_quemimporta[0].value:
                linha_da_coll_quemimporta[0].value = usuario_nome

        try:
            informacao_carteira_object = Carteira.objects.get(cod_cred=cod_cred,nome_cred_padrao=nome_carteira,ativa=True) 
        except:
            messages.add_message(request,messages.ERROR, 'A carteira informada está inativa ou é inexistente. Caso exista, contate-nos para adicionarmos esta carteira')
            return render(request, 'planilhas/metas.html',context)


        palavras_chaves_carteira    =   informacao_carteira_object.palavras_chaves

        # formatando palavras chaves
        while True:
            palavras_chaves_carteira = palavras_chaves_carteira.replace('-',' ')

            if not '-' in palavras_chaves_carteira:
                break
        cod_cred_carteira           =   informacao_carteira_object.cod_cred
        palavras_chaves_carteira    = str(palavras_chaves_carteira).split()
        centro_custo_carteira       =   informacao_carteira_object.centro_custo
        nome_cred_padrao            =   informacao_carteira_object.nome_cred_padrao  


        for linha_looping_df in meta_sheet_dados.iter_rows(min_col=4, max_col=6, min_row=4):
            cod_cred_planilha = str(linha_looping_df[0].value)
            nome_credor_planilha = str(linha_looping_df[1].value).lower().strip()

            if cod_cred_planilha == 'None' or cod_cred_planilha == None:
                break

            if cod_cred_planilha != cod_cred_carteira:
                messages.add_message(request,messages.WARNING, 'Você selecionou a carteira incorreta... O cod cred selecionado é diferente da planiha')
                return render(request, 'planilhas/metas.html',context)

            for palavra in palavras_chaves_carteira:
                if palavra not in nome_credor_planilha:
                    messages.add_message(request,messages.WARNING, 'Você selecionou a carteira incorreta... O nome do credor selecionado é diferente da planiha')
                    return render(request, 'planilhas/metas.html',context)

            linha_looping_df[2].value = informacao_carteira_object.centro_custo
            linha_looping_df[1].value = nome_cred_padrao

         # Validando código Funcionário
        query_script =  utils_deflatores.CODIGO_FUNCIONARIO

        itens = incremento_functions.execute_consulta(query_script)

        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            query = itens[0].fetchall() # Confirmando as alterações
            incremento_functions.desativando(itens[0],itens[1]) # Desativando

        # Formatando query de códigos de func
        range_query = len(query)
        for resultado in range(range_query):
            
            # Formatando conjuntos de dados da query
            cod_funcionario = str(query[resultado][0]).strip()
            cpf_funcionario = str(query[resultado][1]).strip()
            
            # Inserindo conjunto de dados formatados
            query[resultado][0] = cod_funcionario
            query[resultado][1] = cpf_funcionario

        # Entrando em cada código funcionário da planilha
        for celula in meta_sheet_dados.iter_rows(min_row=4,min_col=7,max_col=8):

            cod_func = str(celula[0].value)
            cpf_func = str(celula[1].value)
            cpf_func = formatar_cpf(cpf_func)

            # Buscando o código do funcionário 
            for conjunto_dados in query:
                
                # Pegando o código de funcionário existente no banco de dados
                cod_func_query = conjunto_dados[0]

                # Verificando se o cod func da planilha existe no banco de dados
                if cod_func == cod_func_query:
                    cod_func_correto = True
                    print('O código está correto')
                    break

                else:
                    cod_func_correto = False
            
            if not cod_func_correto:
                
                # Buscando o cpf do funcionário
                for conjunto_dados in query:
                    
                    # Pegando o cpf de funcionário existente no banco de dados
                    cod_func_query = conjunto_dados[0]
                    cpf_func_query = conjunto_dados[1]

                    # Verificando se o cpf func da planilha existe no banco de dados
                    if cpf_func == cpf_func_query:

                        # Inserindo o codigo funcionário correto
                        celula[0].value = cod_func_query

                        break

        for looping_jk in meta_sheet_dados.iter_rows(min_col=10,max_col=11, min_row=4):
            supervisor = looping_jk[0].value
            frente = looping_jk[1].value

            if not frente and not supervisor:
                continue

            if  not type(frente) == str:  
                messages.add_message(request,messages.WARNING, 'Não se pode preencher uma célula apenas com numeros na coluna de FRENTEs')  
                return render(request, 'planilhas/metas.html',context)
            if not type(supervisor) == str:
                messages.add_message(request,messages.WARNING, 'Não se pode preencher uma célula apenas com numeros na coluna de Supervisores')  
                return render(request, 'planilhas/metas.html',context)

        for looping_lo in meta_sheet_dados.iter_rows(min_col=12,max_col=15, min_row=4):
            meta_qtd,meta_honorarios,meta_repasse,meta_valor = looping_lo[0].value, looping_lo[1].value, looping_lo[2].value,looping_lo[3].value

            if not meta_qtd or not meta_honorarios or not meta_repasse or not meta_valor:
                continue

            meta_incorreta = True if type(meta_qtd) == str or type(meta_honorarios) == str or type(meta_repasse) == str or type(meta_valor) == str else False
            if meta_incorreta:
                messages.add_message(request,messages.WARNING, 'As metas devem totalmente numéricas, não pode haver caracteres')  
                return render(request, 'planilhas/metas.html',context)
            
        for looping_q in meta_sheet_dados.iter_rows(min_row=4, max_col=17, min_col=17):
            turno = str(looping_q[0].value).lower().strip()
            
            if turno == 'none':
                break


            turno_possibilidades = (
                'integral',
                'manha',
                'manhã',
                'tarde',
            )

            turno_incorreto = False if turno not in turno_possibilidades else True
            if not turno_incorreto:
                messages.add_message(request,messages.WARNING, 'Há erros na coluna de turnos')  
                return render(request, 'planilhas/metas.html',context)

        for looping_r in meta_sheet_dados.iter_rows(min_row=4, max_col=18, min_col=18):
            atuacao = looping_r[0].value

            if atuacao == None or str(atuacao).strip == 'None':
                break


            if type(atuacao) != str:
                messages.add_message(request,messages.WARNING, 'Há erros na coluna de atuação')  
                return render(request, 'planilhas/metas.html',context)                
 
        for looping_s in meta_sheet_dados.iter_rows(min_row=4, max_col=19, min_col=19):
            estagio = str(looping_s[0].value).lower().strip()
            
            if estagio == 'none':
                break

            estagio_possibilidades = (
                'estágio',
                'estagio',
                'amigável',
                'amigavel',
            )
            if estagio not in estagio_possibilidades:
                messages.add_message(request,messages.WARNING, 'Há erros na coluna de estagios')  
                return render(request, 'planilhas/metas.html',context)                
 

        contador_linha = 4

        colunas = {
            'META_QTDE'        :       'L',
            'META_HONORARIOS'  :       'M',
            'META_REPASSE'     :       'N',
            'META_VALOR'       :       'O',
        }
        for lopping_v in meta_sheet_dados.iter_rows(min_row=4,max_col=22,min_col=22):
            
            tipo_medicao = str(lopping_v[0].value).upper().strip()
            
            if tipo_medicao == 'NONE':
                break
            
            if not tipo_medicao in colunas:
                print(tipo_medicao)
                messages.add_message(request,messages.WARNING, 'O tipo de medição está incorreto.')  
                return render(request, 'planilhas/metas.html',context) 
              
            coluna_informada = colunas[tipo_medicao]
            linha_atual = str(contador_linha)

            celula_atual = coluna_informada + linha_atual
            celula_atual = meta_sheet_dados[celula_atual].value

            if type(celula_atual) != int:
                messages.add_message(request,messages.WARNING, 'Apenas se pode inserir inteiros nas metas!')  
                return render(request, 'planilhas/metas.html',context)   
                
            celula_correta = True if celula_atual and celula_atual != 0 else False
            if not celula_correta:
                messages.add_message(request,messages.WARNING, 'Você preencheu a coluna de tipo_medicao com um valor e inseriu o valor da meta em outra...')  
                return render(request, 'planilhas/metas.html',context)                  


            
            celulas = (
                'L' + linha_atual,
                'M' + linha_atual,
                'N' + linha_atual,
                'O' + linha_atual,
            )
            valores_preenchidos = 0
            for celula in celulas:

                valor_celula = meta_sheet_dados[celula].value
                try:
                    if int(valor_celula) != 0:
                        valores_preenchidos += 1
                except:
                    valores_preenchidos += 1
                    

            contador_linha += 1

            if valores_preenchidos != 1:
                messages.add_message(request,messages.WARNING, 'Você pode preencher apenas um valor das quatro colunas de meta.')  
                return render(request, 'planilhas/metas.html',context)   

        competencia = list(competencias)[0]
        
        ano = competencia[0:4]
        mes = competencia[5:]


        data_comeco = f'{mes}-01-{ano}'
        data_final  = f'{mes}-28-{ano}'

        for looping_tu in meta_sheet_dados.iter_rows(min_row=4, min_col=20, max_col=21):
            if looping_tu[0].value == None:
                break

            looping_tu[0].value = data_comeco
            looping_tu[1].value = data_final

        
        # Criando consulta 
        quantidade_cpf = 0
        lista_cpfs = set()
        nome_funcs = []

        # Adicionando cpfs na lista de cpfs 
        for looping_h in meta_sheet_dados.iter_rows(min_row=4, max_col=9, min_col=8):
            cpf = str(looping_h[0].value).strip().lower()
            if cpf == 'none':
                break

            nome_func = looping_h[1].value.strip().lower()

            quantidade_cpf += 1

            if len(cpf) == 10:
                cpf = '0' + cpf
                looping_h[0].value = cpf
            elif len(cpf) == 9:
                cpf = '00' + cpf
                looping_h[0].value = cpf
                
            lista_cpfs.add(cpf)
            
            nome_funcs.append(nome_func)

        consulta = utils_op.QUERY_FUNCIONARIOS
        cpfs_iterados = 1

        # Passando cada cpf para a consulta, assim formando uma query
        for cpf in lista_cpfs:
            
            if cpfs_iterados == len(lista_cpfs):
                consulta += f"'{cpf}'"
            else:
                consulta += f"'{cpf}',"

            cpfs_iterados += 1
        
        consulta += ')'
        itens =  incremento_functions.execute_consulta(consulta)

        # Verificando se a importação funcionou e caso funcione, vamos recuperar as informações da query
        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            cpfs = itens[0].fetchall() # Confirmando as alterações    
            incremento_functions.desativando(itens[0],itens[1]) # Desativando

        cpfs_funcionarios_query = set()
        nome_funcionarios_query = list()

        # Passando os valores(nome e cpf) da consulta para as estruturas de dados acima
        for valor in cpfs:
            nome    =   valor[0].strip().lower()
            cpf     =   valor[1].strip().lower()
            

            cpfs_funcionarios_query.add(cpf)
            nome_funcionarios_query.append(nome)

        for cpfs in lista_cpfs:
            if cpfs not in cpfs_funcionarios_query:
        
                messages.add_message(request,messages.WARNING, 'Há cpfs de funcionários que estão incorretos em sua planilha')
                return render(request, 'planilhas/metas.html',context)


        for nome in nome_funcs:
            if nome not in nome_funcionarios_query:
                messages.add_message(request,messages.WARNING, 'Há nomes de funcionários que estão incorretos em sua planilha')
                return render(request, 'planilhas/metas.html',context)
        
        # Preenchendo a coluna meta ativa
        for celula in meta_sheet_dados.iter_rows(min_row=4,min_col=16, max_col=16):
            valor_celula = celula[0].value

            if valor_celula == None:
                break

            celula[0].value = 'SIM'
        
        query_metaoperadores =  utils_op.QUERY_META
        where_metaoperadores =  f"COMPETENCIA = '{competencia}' AND "
        where_metaoperadores += f"NOME_CREDOR = '{nome_cred_padrao}' "

        query_metaoperadores += where_metaoperadores       

        itens = incremento_functions.execute_consulta(query_metaoperadores)
        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            query_meta = itens[0].fetchall() # Confirmando as alterações    
            incremento_functions.desativando(itens[0],itens[1]) # Desativando
 
        if len(query_meta) != 0:
            update_metaoperadors = utils_op.UPDATE_META
            update_metaoperadors += where_metaoperadores

            itens = incremento_functions.execute_consulta(update_metaoperadors)
            
            importacao_nao_funcionou = itens[2] 
            if importacao_nao_funcionou:
                incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
                messages.add_message(
                    request,messages.
                    ERROR,
                    "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos alterar meta antiga no banco de dados"
                    )
                return render(request, 'planilhas/metas.html',context)
            else:
                itens[1].commit() # Confirmando as alterações    
                incremento_functions.desativando(itens[0],itens[1]) # Desativando 

        dados = incremento_functions.get_data_meta_com_duplicatas(meta_sheet_dados,max_col=22)
        valores_insert = incremento_functions.criando_valors_para_insert(dados,22)
        insert_consulta = utils_op.INSERT_META
        consulta = insert_consulta + valores_insert


        itens = incremento_functions.execute_consulta(consulta)
        importacao_nao_funcionou = itens[2] 
        if importacao_nao_funcionou:
            incremento_functions.desativando(itens[0],itens[1]) # Desativando a conexão sql
            
            messages.add_message(
                request,messages.
                ERROR,
                "Ops, algo deu errado. Contate o desenvolvedor, não conseguimos importar sua planilha para o banco de dados"
                )
            return render(request, 'planilhas/metas.html',context)
        else:
            try:
                query_meta = itens[0].commit() # Confirmando as alterações    
            except:
                incremento_functions.desativando(itens[0],itens[1]) # Desativando

        messages.add_message(request,messages.SUCCESS,'Meta importada com sucesso!')
        return render(request, 'planilhas/metas.html',context)
    
    return render(request, 'planilhas/metas.html',context)# TODO: Criando context -- meta de operadores -- alterar
